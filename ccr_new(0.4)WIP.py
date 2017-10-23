#!/usr/bin/evn python

"""
Selenium script that enters details into CCR based on data read from
an Excel spreadsheet.

0.1 - initial icomplete version that uses ccr_data(0.1).xlsx
0.2 - changed to use different spreadsheet format
0.3 - Removed basic_fee and calc_fee calls from CCR.create_claim
        Modified CCR.basic_fee() to handle disabled fields. Shebang line added!
0.4 - Moved Excel Management into new class ExcelRun
"""

#Standard Selenium imports
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoAlertPresentException, NoSuchElementException
from selenium.webdriver.support.ui import Select, WebDriverWait

#Needed to set particular Firefox profile - used to login using modify headers
from selenium.webdriver.firefox.webdriver import FirefoxProfile

#Used to read spreadsheet
import openpyxl

#Generally useful
import os, time, datetime, sys

class ExcelRun:
    def __init__(self, filename=""):
        """
        Uses Selenium to carry out various CCR activities based on data
        data from specially formatted spredsheet.
        Args:
            filename - Excel file with test data
        """
        #Excel filename
        self.filename = filename
        #CCR webdriver class
        self.ccr = CCR()
        #Run tests using data from self.filename spreadsheet if supplied
        if filename:
            self.excel_run()

    def excel_run(self, endrow = 20):
        """Run using data from spreadsheet
        Args:
            endrow (int) - maximum row number when reading scenarios from
                           run tab. Optional - defaults to 20
        """
        #Try to open spreadsheet
        try:
            self.wb = openpyxl.load_workbook(filename=self.filename)
        #Give up if fails
        except Exception as e:
            response = self.filename+" - Failed to read: "+e.__doc__
            print response
            print os.getcwd()
            #Give up if unsuccessful
            return response

        #Check that required tabs are present and give up if any missing
        essential_tabs = ["Run","Login"]
        missing_tabs = [tab for tab in essential_tabs if tab not in self.wb.sheetnames]
        if missing_tabs:
            print "Excel file:", self.filename
            print "Ending - required tab(s) missing: "+ ", ".join(missing_tabs)
            return "Missing tabs: "+", ".join(missing_tabs)

        #Setup test results filename and folder
        results_folder = os.path.join(os.getcwd(),"ccr_results")
        #Create the folder if it doesn't exist
        if not os.path.exists(results_folder):
            os.makedirs(results_folder)
        #Results filename - based or source filename but with date/time added
        results_filename = (os.path.splitext(self.filename)[0]
                            +time.strftime("_results_[%Y.%m.%d_%H.%M.%S].xlsx"))
        results_file = os.path.join(results_folder, results_filename)

        #Read defined test data from spreadsheet
        login_data = self.excel_data_read("Login")
        claim_data = self.excel_data_read("Claim")
        basic_data = self.excel_data_read("Basic")
        expen_data = self.excel_data_read("Expenses")

        #Execute each test scenario from the Run tab
        ws = self.wb.get_sheet_by_name("Run")
        #Find positions of each expected column heading
        columns = self.excel_column_positions("Run", heading_row=6, leftcol=1, maxcol=20)
        #Process each row from spreadsheet
        for row in range(7, endrow+1):
            print "Processing row:", row
            #Skip to next row (scenario) if skip set
            skip = ws.cell(row=row, column=3).value
            if skip:
                print "Skipped - skip flag set."
                continue
            #Create tab for recording results for the row
            results_name = "Results_"+str(row)
            self.wb.create_sheet(title=results_name)
            #Results sheet
            rs = self.wb.get_sheet_by_name(results_name)
            #Text at top
            rs.sheet_view.showGridLines = False
            rs["A1"].value = "Tab added by script on "+time.strftime("%d-%b-%Y %H:%M:%S")
            rs["A1"].font = openpyxl.styles.Font(bold=True)
            result_row = 3

            #Login part of scenario
            login_key = ws.cell(row=row, column=columns["login"]).value
            login_key = str(login_key).lower()
            ##print row, login_key
            if login_key !="none":
                login_args = login_data.get(login_key,"")
                #Try to login if login arguments supplied
                if login_args:
                    message = "URL: "+login_args.get("url","")+", User: "+login_args.get("username","")
                    ##print login_args
                    ok = self.ccr.login(**login_args)
                    if ok:
                        message = message +" - Login sucessful."
                    #skip to next row if login unsuccessful
                    else:
                        message = message + "- Login failed. Skipping to next row."
                else:
                    message = "No login arguments associated with label, "+login_key+". Skipping to next row."
            else:
                message = "No login arguments specified. Skipping to next row."
            #Record and display message
            rs.cell(row = result_row, column = 1).value = message
            print message
            #Skip to next row if login was not possible/unsuccessful
            if message.endswith("Skipping to next row."):
                continue

            #Get CCR version from screen and write to spreadsheet
            result_row += 1
            rs.cell(row = result_row, column = 1).value = "Version: "+self.ccr.get_version()

            #Claim part of the scenario
            message = "Claim not completed."#defult unsuccessful message
            claim_key = ws.cell(row=row, column=columns["claim"]).value
            claim_key = str(claim_key).lower()
            if claim_key !="none":
                claim_args = claim_data.get(claim_key,"")
                if claim_args:
                    #Need to convert defendants from csv to list
                    defendants = claim_args.get("defendants","")
                    defendants = [c.strip() for c in str(defendants).split(",")]
                    claim_args["defendants"] = defendants
                    ##print "Claim details:",claim_args
                    self.ccr.create_claim(**claim_args)
                    message = "Claim completed using : "+claim_key
            #Write results to results tab
            result_row += 1
            rs.cell(row = result_row, column = 1).value = message

            #Basic Fee part of scenario
            message = "Basic fee not entered"
            basic_key = ws.cell(row=row, column=columns["basic"]).value
            basic_key = str(basic_key).lower()
            if basic_key !="none":
                basic_args = basic_data.get(basic_key,"")
                if basic_args:
                    self.ccr.basic_fee(**basic_args)
                    message = "Basic fee:"

            # Expenses part of scenario
            # Treat as comma-separated values
            message = "Expenses not entered"
            expen_keys = ws.cell(row=row, column=columns["expenses"]).value
            if expen_keys:
                #Extract comma-separated values also convert to lower-case and strip spaces
                expen_keys = [str(key).lower().strip() for key in expen_keys.split(",")]
                #Enter details from each expenses key
                #NEED TO ADD HANDLING FOR PRESSING NEW ADD CLAIM ELEMENT BUTTON
                rowpos = 0#row position for entering expenses on screen
                for expen_key in expen_keys:
                    if expen_key !="none":
                        expen_args = expen_data.get(expen_key,"")
                        if expen_args:
                            self.ccr.expenses(row=rowpos, **expen_args)
                            rowpos+=1
                            message = "Expenses:"

            #Calculate Fee
            self.ccr.calc_fee()

            #Read some figures from CCR front-end
            figures = self.ccr.read_fees()
            print figures
            #Write results to results tab
            result_row += 1
            rs.cell(row = result_row, column = 1).value = message
            for dc, (heading, val) in enumerate(figures.iteritems()):
                rs.cell(row = result_row,   column = 1+1+dc).value = heading
                rs.cell(row = result_row+1, column = 1+1+dc).value = val
            result_row += 1

        #Add yellow stripe to top of each tab, to make results spreadsheet
        #distinctive
        fill = openpyxl.styles.PatternFill(start_color='50FFFF00',
                                            end_color='50888800',
                                            fill_type='solid')
        for ws in self.wb.worksheets:
            for column in range(1, 30):
                ws.cell(row=1,column=column).fill = fill

        self.wb.save(results_file)
        print "Results saved to:", results_filename
        #End

    def excel_data_read(self, tab, toprow=6, maxrow=20, leftcol=3, maxcol=20):
        """Reads test data in from spreadsheet tab
        Args:
            tab - name of spreadsheet tab
            toprow - first row to read (must contain keys)
            maxrow - last row to read
            leftcol - leftmost column to read from
            maxcol - last column to read from
        Returns:
            dictionary of dictionaries containing the data. Key in
            outermost dict is taken from the tab's "Label" column. Inner
            dictionaries hold data from each row.
            All Keys are converted to lower-case strings
        """
        #Select tab
        ws = self.wb.get_sheet_by_name(tab)

        #Read headings (will be used as dictionary keys)
        #and map to its column number to make reading easier
        heading_columns = self.excel_column_positions(tab, toprow, leftcol, maxcol+1)

        #Read the data into a dictionary of dictionaries
        tab_data = {}
        for row in range(toprow+1, maxrow+1):
            #Get key for outer dictionary (tab_data) from the Label column
            label_column = heading_columns["label"]# remember converted to lower-case, so "label"
            key = ws.cell(row=row, column=label_column).value
            key = str(key).lower()#ensure lower-case
            #default key to row number if value blank (blank is "none" at this point due to str().lower() above)
            if key =="none":
                key = str(row)
            #Get the values from the columns other than "label"
            temp = {}
            for heading, column in heading_columns.iteritems():
                if heading != "label":
                    val = ws.cell(row=row, column=column).value

                    #Special value generation using date_maker function
                    if type(val) in (str, unicode):
                        if val.startswith("#"):
                            val = data_maker(val)

                    #Some reformatting of data
                    #Ensure date are d/m/y strings
                    if type(val) is datetime.datetime:
                        val = val.strftime('%d/%m/%Y')
                    #Convert numbers to str
                    if type(val) in (int, float):
                        val = str(val)
                    #Store the value in inner dictionary, using column heading as key
                    temp[heading]=val

            #Assign the row data dict to the key in the containing dict
            tab_data[key]=temp
        return tab_data

    def excel_run_col(self, tab, column):
        """
        OBSOLETE - no longer used but like the way it works!

        Run test using a column of test data
        Args:
            tab: name of tab with data
            column: number of column with data"""
        startrow = 5
        ws = self.wb.get_sheet_by_name(tab)
        run_data={}
        #Read data from spreadsheet
        for row in range(startrow, 200):
            key = ws.cell(row=row, column=2).value
            if key:
                val = ws.cell(column+str(row)).value
                #Convert dates to d/m/yyyy format strings
                ##print val,type(val)
                if type(val) is datetime.datetime:
                    val = val.strftime('%d/%m/%Y')

                #Convert numbers to str
                if type(val) in (int, float):
                    val = str(val)

                run_data[key]=val

        #Reprocess certain items
        #Convert defendants from csv to list
        if "defendants" in run_data:
            run_data["defendants"]=run_data["defendants"].split(",")

        #Create claim using data read from spreadsheet
        ##print run_data
        ##self.create_claim(**run_data)

    def excel_column_positions(self, tab, heading_row=6, leftcol=3, maxcol=20):
        """Finds the positions of columns in spreadsheet bases on unique
        text labels in a heading row.

        Args:
            tab - Name of tab on which to look
            heading_row - row number in which to look
            leftcol - first column to examine (as number)
            maxcol - last column to examine (as number)

        Returns:
            dirctionary with text labels as keys and column numbers as values
            e.g. {"Name":1, "Date":2, "ID":3 }
        """
        #Select tab
        ws = self.wb.get_sheet_by_name(tab)
        #Read headings (will be used as dictionary keys)
        #and map to its column number to make reading easier
        heading_columns = {}
        for column in range(leftcol, maxcol+1):
            val = ws.cell(row=heading_row, column=column).value
            if val:
                val = str(val).lower()#ensure lower-case
                heading_columns[val]=column
        return heading_columns

class CCR:
    def __init__(self):
        """
        Uses Selenium to log-in to CCR and create cases.
        """
        # #Holds names tabs present on the CCR screen. Set by self.ccr_tabs
        self.tabs = []
        ##self.direct_run()

    def direct_run(self, url, username = "", password = "", ffp = ""):
        """Directly interact with CCR, not using spreadsheet data.
        Mainly used to try things out
        """
        login_response = self.login(url, username, password, ffp)
        if login_response:
        #self.search(caseno="T20140001")
        ##self.ccr_tabs()
            self.create_claim()
        #self.logout()

    def login(self, url, username="", password="", firefox_profile=""):
        """
        Login to CCR either via Portal or Modify Headers(using Firefox profile)
        Args:
            url - url (either portal or /ccr/Autlogin for Modify headers)
            username - portal username (not needed if Modify headers being used)
            passwored - portal password  (not needed if Modify headers being used)
            firefox_profile - path to firefox profile (only needed if Modify headers being used)

        # *NEEDS ADAPTING FOR NEW PORTAL*

        Returns:
            True if login succesfull, otherwise False
        """
        #Create webdriver instance.
        #Diferent settings depending on whether we're accessing via Modify Headers
        if firefox_profile:
            #Use specified Firefox Profile (Modify Headers)
            ffp_object = FirefoxProfile(firefox_profile)
            self.driver = webdriver.Firefox(ffp_object)
        else:
            #Normal default Firefox profile (portal login)
            self.driver = webdriver.Firefox()
        driver = self.driver

        result = False #flag will be changed to True if login successful
        if firefox_profile:
            #Direct login (modify headers)
            driver.get(url)
        else:
            #Normal Portal Login (old portal)
            driver.get(url)
            driver.find_element_by_id("userField").send_keys(username)
            driver.find_element_by_id("passField").send_keys(password)
            driver.find_element_by_name("submit").click()

            #Wait a bit for response
            try:
                WebDriverWait(self.driver,10).until(lambda driver:
                    ">Logged in as:" in driver.page_source
                    or ">Authentication failed" in driver.page_source
                    or "This page can?t be displayed" in driver.page_source
                    or "Secure Connection Failed" in driver.page_source
                    or "HTTP Status 404" in driver.page_source)
            except TimeoutException:
                pass

            #Stop is login was not successful
            if "Logged in as:" not in driver.page_source:
                return False

            ##print  [e.text for e in driver.find_elements_by_tag_name("a")]
            #Click portal link
            link = u'Crown Court Remuneration (CCR)'
            #Set new_win to be "y" if link's target attribute is 'LSCAppPopUp'
            new_win = False
            if driver.find_element_by_link_text(link).get_attribute("target") == 'LSCAppPopUp':
                new_win = True
            driver.find_element_by_link_text(link).click()

            # Need to switch to new window
            if new_win:
                ##WebDriverWait(driver,20).until(lambda x: len(x.window_handles)==2,self.driver)
                WebDriverWait(driver,20).until(lambda x: len(x.window_handles)==2)
                newwindow = driver.window_handles[-1]
                driver.switch_to_window(newwindow)

        #Wait for expected CCR page
        WebDriverWait(self.driver, 20).until(lambda driver:
                                "Search For Claims" in driver.page_source
                                or "<h2>Login Error</h2>" in driver.page_source)

        #Set success flag if login successful
        if "Search For Claims" in driver.page_source:
            result = True

        return result

    def search(self, caseno="", court="", repno="", supplier=""):
        """Complete search in CCR"""
        driver = self.driver
        #Click search tab
        ##driver.find_element_by_css_selector("span").click()
        self.ccr_tabs(click_link="Search For Claims")

        #Complete search fields
        if caseno:
            driver.find_element_by_id("caseNumber").clear()
            driver.find_element_by_id("caseNumber").send_keys(caseno)#"T20132011"
        if court:
            Select(driver.find_element_by_id("court")).select_by_visible_text(court)#"Basildon (461)"
        if repno:
            driver.find_element_by_id("representationOrderNumber").clear()
            driver.find_element_by_id("representationOrderNumber").send_keys(repno)
        if supplier:
            driver.find_element_by_id("supplierID").clear()
            driver.find_element_by_id("supplierID").send_keys(supplier)

        #Click the search button
        driver.find_element_by_xpath("//input[@value='Search']").click()

        #Wait for sarch results
        WebDriverWait(driver,10).until(lambda driver: "Search Results" in driver.page_source
            or "No claims found" in driver.page_source,driver)

        #If we've results examine them
        if "Search Results" in driver.page_source:
            print "Search has results as follows:"
            #Loop to move through any multi-page results
            keep_going = True
            while keep_going:
                #Examine dispalyed results
                self.search_results()
                #Is there a "Next" button
                next_buttons = [e for e in driver.find_elements_by_class_name("button") if e.get_attribute("value")==u"Next"]
                ##print "nb", next_buttons
                #If next button click it to advance to next page, otherwise exit loop
                if next_buttons:
                    next_buttons[0].click()
                else:
                    keep_going = False

    def search_results(self):
        """Examine page of CCR search results

        ['T20140001','T20140001\439','1','233JK','MCMILLAN','LINDSAY','19/01/1927','28/11/2013','07/07/2014','N']
        """
        driver = self.driver
        tables = driver.find_elements_by_tag_name(u"table")
        #Results are in 5th table table
        headings = [e.text for e in tables[4].find_elements_by_tag_name('th')]
        print "Headings",headings
        #List data for each row
        rows = tables[4].find_elements_by_class_name(u"dataRowo")
        for row in rows:
            #Get the row text
            row_text = [e.text for e in row.find_elements_by_tag_name("td")]
            print ",".join(row_text)
            #Find the select button
            buttons = row.find_elements_by_class_name("button")
            #print "bt>", [e.get_attribute("value") for e in buttons]

    def create_claim(self,
                    claim_type="Rep Order Issued in the Crown Court Subject to Means Testing",
                    rep_order_date="01/10/2013",
                    mags_court="Barking (2)",
                    supplier = "02AYC",
                    ccourt = "Oxford (445)",
                    case_no="T20170002",
                    defendants=["3264731", "3264732", "3264733"], #this is the MAATID
                    prosecution = "Crown Prosecution Service (CPS)",
                    first_day = "01/09/2017"
                    ,**kwargs):
        """
        Start new CCR claim.
        Completes initial details
        Args:
            (details to be added!)
            **kwargs included in case surplus arguments supplied, such as
            value read from a spreadsheet
        """
        driver = self.driver

        #See if we're already on the right screen. If not navigate to it.
        self.ccr_tabs()
        current_tab = self.current_tab()
        if current_tab!="Create Claim":
            self.ccr_tabs(click_link="Create Claim")

        #Could be alert message here if there is unsaved data in CCR
        #Check for message and accept it if found.
        self.alertcheck(accept=True)

            ##Wait for Create Claim to become the current tab (_ arg convention of unused/unwanted argument in lambda fn)
            ##WebDriverWait(self.driver,10).until(lambda _: self.current_tab()=="Create Claim")

        #PAH no tabs present on Create Claim when screen first accessed (although appear after some details entered)!

        #Wait for screen
        WebDriverWait(driver,10).until(lambda driver:"Please select type of claim you wish to create and press the" in driver.page_source)

        #Select case type
        Select(driver.find_element_by_id("ccrCaseType")).select_by_visible_text(claim_type)
        #"Create" button
        driver.find_element_by_id("caseTypeGoButton").click()

        #Rep order check thing
        driver.find_element_by_id("repOrderDateForCheck").send_keys(rep_order_date)
        Select(driver.find_element_by_id("magsCourt")).select_by_visible_text(mags_court)
        driver.find_element_by_id("goButton").click()

        #Case Details
        driver.find_element_by_id("accountNumber").send_keys(supplier)
        driver.find_element_by_id("caseNumber").send_keys(case_no)
        Select(driver.find_element_by_id("courtCode")).select_by_visible_text(ccourt)
        Select(driver.find_element_by_id("prosecutingAuthorityCode")).select_by_visible_text(prosecution)

        #Add Claim Elements - can get stuck here if indictment number not unique
        #within some other things
        while True:
            driver.find_element_by_id("bAddClaimElements").click()

            #Error check
            errors = self.error_check()
            if errors:
                print "".join(errors)

            #Indictment number needs to be unique (in combination with something else)
            #Increment the number if error message displayed
            if "The entered case already exists. Please increment the Indictment Number (for a Severed Case or Solicitor Advocate claim) or enter as a Re-trial" in errors:
                indictment_no = driver.find_element_by_id("indictmentNumber")
                value = indictment_no.get_attribute("value")
                value = str(int(value)+1)
                indictment_no.clear()
                indictment_no.send_keys(value)
            else:
                break

        #Trial details
        Select(driver.find_element_by_id("offenceClass")).select_by_visible_text("A Homicide and related grave offences")
        Select(driver.find_element_by_id("offenceCode")).select_by_visible_text("Soliciting to commit murder")
        Select(driver.find_element_by_id("personType")).select_by_visible_text("QC")
        Select(driver.find_element_by_id("scenario")).select_by_visible_text("Trial")
        driver.find_element_by_id("estTrialLength").send_keys("7")
        driver.find_element_by_id("trialLength").send_keys("9")
        driver.find_element_by_id("firstDayOfTrial").send_keys(first_day)

        #Add defendants
        self.add_defendants(defendants)


    def add_defendants(self,maat_ids=["3264731"]):
        """Add one or more defendants to CCR claim based on supplied MAAT IDs
        Create claim screen must be open with Add Defendants button present
        (after "Add Claim Elements" pressed)
        On return to main claim screen sets first defendant to be "Main Def"
        if none already ticked.
        Args:
            maat_ids - list of maat IDs to be used in search for
                e.g. ["3264731","3264732","3264733"]
        """
        driver = self.driver

        #Press "Add Defendants" button
        driver.find_element_by_id("bAddDefendants").click()

        #Search for each supplied MAAT ID in turn
        for maat_id in maat_ids:
            WebDriverWait(driver,10).until(lambda driver: len(driver.find_elements_by_id("maatReference"))>0)
            driver.find_element_by_id("maatReference").clear()
            driver.find_element_by_id("maatReference").send_keys(maat_id)
            print "m", maat_id

            #Click Search button
            driver.find_element_by_xpath("//input[@value='Search']").click()

            #Wait for defendant search result
            WebDriverWait(driver,10).until(lambda driver: "Search Results" in driver.page_source
                or "No defendants found" in driver.page_source)

            # "No defendants found, please refine the search criteria and try again"
            if "No defendants found, please refine the search criteria and try again" in driver.page_source:
                pass
            else:
                driver.find_element_by_id("selectDefendant[0]").click()
                driver.find_element_by_xpath("//input[@value='Add to Case']").click()

        #Return to main claim screen
        driver.find_element_by_css_selector("#container > table > tbody > tr > th.left_border > input.button").click()

        #Select one defendant as the main one if it's none already ticked
        table = driver.find_element_by_id("defendantsTable")
        ##first_defendant = driver.find_element_by_id("mainDefendant[0]")
        #Find defendants (css_selector used to get partial matching via id)
        defendants = table.find_elements_by_css_selector("input[id*='mainDefendant']")
        #If we have defendants make sure one is ticked as the main one.
        if defendants:
            if True not in [e.is_selected() for e in defendants]:
                defendants[0].click()

    def basic_fee(self,
                claim_element="Advocate Fee",
                ppe="1",
                attendance="1",
                witnesses="0",
                defendants="", cases="",**kwargs):
        """Enter Basic Fee details
        Relevant section of the claim screen must be available for this to work
        Args:
            claim_element - Claim Element value
        """
        driver = self.driver
        #Basic fee fields are within this HTML table
        #Element ids are unique within table but could be duplicated within
        #page, so items selected at table-level
        table = driver.find_element_by_id("basicFeeTable")

        #Claim element drop-down
        Select(table.find_element_by_id("billSubType")).select_by_visible_text(claim_element)

        #Input fields are locked until Claim Element selected, so wait for
        #them to become available (look for style="display: none" attribute)
        WebDriverWait(self.driver,10).until(lambda driver:table.find_element_by_id("ppe").get_attribute("style")!="display: none")

        #Complete the text fields
        #Map field element ids to arguments
        field_mapping =[
        ("ppe", ppe), # PPE
        ("numberOfAttendanceDays", attendance), # No. of Attendance Days
        ("numberOfWitnesses", witnesses), # No. of prosecution witnesses
        ("defendants", defendants), # No. of defendants (usually set automatically)
        ("cases", cases), # No of cases (usually set automatically)
        ]

        #Update each in turn
        #Below fairly generic - could move to dedicated method (Webdriver WebElement)
        for elem_id, value in field_mapping:
            #Only update if value supplied (as may sometimes want to leave
            #original value in place)
            if value:
                field = table.find_element_by_id(elem_id)
                #Don't try to update if the field is not on display
                #(Not sure where terminal ";" comes from! Didn't see it in source!)
                if field.get_attribute("style")!="display: none;":
                    ##print ">>",elem_id, field.get_attribute("style")
                    table.find_element_by_id(elem_id).clear()
                    table.find_element_by_id(elem_id).send_keys(value)
                else:
                    print "Field:", elem_id, "-not on display."

        #Original old way
        """
        #PPE
        table.find_element_by_id("ppe").clear()
        table.find_element_by_id("ppe").send_keys(ppe)

        #Number of Attendance Days
        table.find_element_by_id("numberOfAttendanceDays").clear()
        table.find_element_by_id("numberOfAttendanceDays").send_keys(attendance)

        #No of Prosecuting Witnesses
        table.find_element_by_id("numberOfWitnesses").clear()
        table.find_element_by_id("numberOfWitnesses").send_keys(witnesses)

        #No. of Defendants. Automatically set but amend if value supplied
        if defendants:
            table.find_element_by_id("defendants").clear()
            table.find_element_by_id("defendants").send_keys(defendants)

        #No. of Cases. Automatically set but amend if value supplied
        if cases:
            table.find_element_by_id("cases").clear()
            table.find_element_by_id("cases").send_keys(cases)
        """


    def multi_element_set(self):
        """Adds/removes multi-element rows
        Some fee types can have multiple-rows
        Generic - used by self.misc_fee, self.expenses
        """
        pass


    def misc_fee(self,
                claim_element,
                ref_no,
                occurrence_date,
                defendants,
                quantity
                ):
        """Completes Miscellaneous Fees part of claim

        """
        driver = self.driver

        #Expenses fields are within this HTML table
        table = driver.find_element_by_id("miscFeeTable")
        #Can be multiple rows. Need right one for this item
        body = table.find_element_by_tag_name("tbody")
        brows = body.find_elements_by_tag_name("tr")



    def expenses(self,
                claim_element = "Conferences & Views - Car",
                date_incurred = "01/10/2017",
                description = "things",
                vat_rate = "",
                quantity = "1",
                rate = "1.23",
                row = 0
                ):

        """Complete expense - under construction!
        May needs updating because id's shift when new expenses row added

        Args:
            row: claim element row number
        """
        driver = self.driver

        #Warning these fields are troublesome
        #Can have multiple rows which when used cause element ids to change
        #Might be better to rely on position in HTML table

        #Expenses fields are within this HTML table
        table = driver.find_element_by_id("agfsExpensesTable")
        #Can be multiple rows. Need right one for this item
        body = table.find_element_by_tag_name("tbody")
        brows = body.find_elements_by_tag_name("tr")

        #If required row does not exist, add it
        rowcount = len(brows)
        print "expenses row thing:",row, rowcount
        if 1+row > rowcount:
            deficiency = 1+row - rowcount
            for _ in range(deficiency):
                #"Add claim element" button in Expenses part of screen
                driver.find_element_by_xpath("(//input[@value='Add Claim Element'])[3]").click()
            #Refresh the list of rows
            brows = body.find_elements_by_tag_name("tr")

        #Set row to one wanted
        row = brows[row]

        #Claim Element
        Select(row.find_element_by_id("billSubType")).select_by_visible_text(claim_element)

        #Date Incurred
        # uses partial matching using css_selector because there's
        # number on end of the id that changes
        field = row.find_element_by_css_selector("input[id*='dateIncurred']")
        field.clear()
        field.send_keys(date_incurred)
        ##row.find_element_by_id("dateIncurred6").clear()
        ##row.find_element_by_id("dateIncurred6").send_keys(date_incurred)

        #Description
        row.find_element_by_id("description").clear()
        row.find_element_by_id("description").send_keys(description)
        #Quantity
        row.find_element_by_id("quantity").clear()
        row.find_element_by_id("quantity").send_keys(quantity)
        #Rate
        row.find_element_by_id("rate").clear()
        row.find_element_by_id("rate").send_keys(rate)


    def calc_fee(self):
        """Presses Calculate Fee button
        The Claim Screen must be already open for this to work
        """
        driver = self.driver
        driver.find_element_by_xpath("//input[@value='Calculate Fee']").click()

        #Wait for please wait message to go - doesn't work
        ##WebDriverWait(self.driver,10).until(lambda driver: '<div id="pleaseWait"><span>Please wait...</span></div>' not in driver.page_source)
        ##self.save_page_source()


    def read_fees(self):
        """Reads some fee figures
        Returns a dictionary containing some figures.
        Also gets the totals from bottom of screen
        """
        driver = self.driver

        #Find all the tables on the page
        """
        tables = driver.find_elements_by_tag_name("table")
        print ">>>>>>>>>>>>>>>>>>>>"
        print tables
        table_ids = [t.get_attribute("id") for t in tables]
        print table_ids
        """
        #Read Basic Fee
        #Warning several fields share the same ID e.g. "agfsFeeTotal"
        table = driver.find_element_by_id("basicFeeTable")
        rpa = table.find_element_by_id("requestedPaymentAmount").get_attribute("value")
        #looks like there each "agfsFeeTotal" figure is duplicated with a hidden and displayed version of each
        #so need to filter out one of each pair
        others =    [e.get_attribute("value")
                    for e in table.find_elements_by_id("agfsFeeTotal")
                    if e.get_attribute("type")!="hidden"]
        figures = [rpa] + others
        #Store the  figures in a dictionary
        figures ={}
        figures["Basic Ammount 1"] = rpa
        for key, value in zip (["Basic Ammount 2","Basic VAT","Basic Total"],others):
            figures[key]=value
        print figures

        #Overall results
        #Element ids of items that hold the totals at bottom of screen
        calc_results_ids = {"GT Basic":"agfsFeeAmtTotal",
                            "GT Misc":"agfsMiscFeesAmtTotal",
                            "GT Advance":"afgsAdvanceAmtTotal",
                            "GT Expenses":"agfsExpAmtTotal",
                            "GT Correction":"correctionAmtTotal",
                            "GT VAT":"totalAGFSVATAmount",
                            "GT Overall":"overallClaimAmount"}
        #Extract the values from the page
        for key, e_id in calc_results_ids.iteritems():
            figures[key] = driver.find_element_by_id(e_id).get_attribute("value")

        return figures

    def logout(self):
        """Logout from CCR"""
        driver = self.driver

        #Find all (any) Exit buttons on the screen
        exits = driver.find_elements_by_link_text('Exit Crown Court Remuneration')
        #If any found, click the first one
        if exits:
            exits[0].click()
        #Click alert, if there is one
        try:
            alert = driver.switch_to_alert()
            alert.accept()
        except:
            print "no alert to accept"

    def field_finder(self, table_id="basicFeeTable"):
        """Finds input fields within HTML table
        Args:
            table_id - element id of HTML table to be examined e.g. "basicFeeTable"
        """
        driver=self.driver
        table = driver.find_element_by_id(table_id)

        drop_downs = [e.get_attribute("id") for e in table.find_elements_by_tag_name("select")]

        #Find enabled text entry fields
        #Avoid fields with style attribute "display: none" as they're disabled - doesn't work
        fields = [e.get_attribute("id")
                for e in table.find_elements_by_tag_name("input")
                if e.get_attribute("style")!="display: none"]

        print drop_downs
        print fields

    def table_finder(self, filename=""):
        """Gathers details of HTML tables on screen
        Optionally writes details to a file.
        Returns:
            list of dictionaries containing info
        """
        driver = self.driver
        tables = driver.find_elements_by_tag_name("table")
        #Gather info from the HTML tables present
        info = []
        for i, table in enumerate(tables):
            temp = {}
            temp["position"] = i
            ##temp["wd_instance"] = table
            temp["id"] = table.get_attribute("id")
            temp["headings"] = [e.text for e in table.find_elements_by_tag_name("th")]
            rows = table.find_elements_by_tag_name("tr")
            info.append(temp)
        # Write to file if filename set
        if filename:
            with open(filename,"w") as f:
                for row in info:
                    line = "Position: "+str(row["position"])+"\n"
                    line = line +" ID:"+row["id"]+"\n"
                    line = line +" Headings:"+", ".join(row["headings"])+"\n"
                    line = line.encode("utf-8")# to stop UnicodeEncodeError errors
                    f.write(line+"\n")
        return info

    def ccr_tabs(self,click_link=""):
        """Find tabs present on CCR screen. Stores names in self.tabs
        Optionally can be used to select tab if click argument supplied.
        Args:
            click_link - optional link text to click
        """
        driver = self.driver
        tab_holder = driver.find_element_by_id("tabbedMenu")
        #Extract the text from the hyperlinks in the tabbed menu
        self.tabs = [e.text for e in tab_holder.find_elements_by_tag_name("a")]

        #Click hyperlink if supplied and it's present
        if click_link:
            if click_link in self.tabs:
                driver.find_element_by_link_text(click_link).click()
            else:
                print "Cannot click link '"+click_link+"' because it's not present:",self.tabs

    def current_tab(self):
        """Finds currently selected tab.
        Relies on tab names recorded in self.tabs
        Returns:
            Currently selected tab name
        """
        driver = self.driver
        #Find the currently selected tab based on attribute class="selected"
        current = ""
        for tab in self.tabs:
            if driver.find_element_by_link_text(tab).get_attribute("class")=="selected":
                current = tab
                break
        return current

    def error_check(self):
        """Returns any error messges present on screen
        Returns:
            list containing error message text
        """
        driver = self.driver
        #Find any error class elements on screen
        errors = driver.find_elements_by_class_name("errors")
        #Extract text from error li tags within each error
        error_messages = []
        for error in errors:
            error_messages = error_messages + [e.text for e in error.find_elements_by_tag_name("li")]
        return error_messages

    def get_version(self):
        """Reads CCR version from screen
        Returns:
            string of concatenated h2 headings
        """
        driver = self.driver
        h2_headings=[e.text for e in driver.find_elements_by_tag_name("H2")]
        return ", ".join(h2_headings)

    def alertcheck(self, accept=False):
        """Generic - Checks for presence of Alert
        Args:
            accept (bool) - automatically accept the alert if set to true
        Returns:
            True if alert found, otherwise False
        """
        driver=self.driver
        #current_win=self.driver.current_window_handle#find current window because we'll want to switch back to it later
        alert = self.driver.switch_to_alert()#switch_to_alert is buggy. It always finds one even if none is present.
        #Can only actualy tell if alert is present when try to interact with it, hence exception handling below.
        try:
            _ = alert.text
        #Warning - NoAlertPresentException needs to have been imported from selenium.common.exceptions
        except NoAlertPresentException:
            message = False
        else:
            message = True
            #Accept the alert if requested
            if accept:
                alert.accept()
        #driver.switch_to_window(current_win)
        return message

    def save_page_source(self):
        """Saves page source to file (filename automatic, includes date/time)
        """
        filename = time.strftime("page_source_%d-%b-%Y_%H.%M.%S.txt")
        with open(filename,"w") as fout:
            fout.write(driver.page_source.encode("utf-8"))


def data_maker(value_format):
    """Generates convenient values for testing
    based on format of value_format argument.

    Currently just produces relative dates but could be expanded

    Args:
        value_format - string defining format.
            e.g. "#d-100" - 100 days before today's date as d/m/Y
                 "#d+1" or "#d1" - tomorrow's date
    Returns
        value based on format string (if processed without problem) as string
        , otherwise returns None
    """
    #Default return value
    value = None
    #Relative Date maker
    if value_format[:2] in ("#d","#D"):
        try:
            delta = int(value_format[2:])
        except exectpion as e:
            return None
        #Calculate date based on day interval supplied
        today = datetime.date.today()
        day_interval = datetime.timedelta(days=1)
        run_date = today + day_interval*delta
        value = run_date.strftime("%d/%m/%Y")
    return value


if __name__ == "__main__":
    print "Starting"
    #Spreadsheet with data
    filename = "ccr_data(0.2).xlsx"

    #Replace filename with command-line argument if we have any
    #(>1 because first argument is this very Python file so doesn't count)
    if len(sys.argv)>1:
        filename = sys.argv[1:2]
    go = ExcelRun(filename=filename)
    print "Finished"




